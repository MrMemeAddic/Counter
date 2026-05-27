"""
Unit tests for counter.py — covers crossing logic, track pruning,
key handling, report generation, reset, and XML export.
"""

import os
import sys
import pytest
import xml.etree.ElementTree as ET
from unittest.mock import patch, MagicMock
import numpy as np

# Ensure the project directory is on sys.path
sys.path.insert(0, os.path.dirname(__file__))

from counter import (
    get_side,
    detect_crossing,
    prune_stale_tracks,
    handle_key,
    generate_reports,
    generate_xml,
    reset_counter,
    init_capture,
)


# ═══════════════════════════════════════════
# get_side
# ═══════════════════════════════════════════

class TestGetSide:
    """Test the side-determination helper."""

    def test_left_side(self):
        assert get_side(cx=280, line_x=300, buffer=10) == "LEFT"

    def test_right_side(self):
        assert get_side(cx=320, line_x=300, buffer=10) == "RIGHT"

    def test_in_buffer_zone_returns_none(self):
        assert get_side(cx=295, line_x=300, buffer=10) is None

    def test_exactly_at_left_edge(self):
        """cx == line_x - buffer is NOT left (it's in the zone)."""
        assert get_side(cx=290, line_x=300, buffer=10) is None

    def test_exactly_at_right_edge(self):
        """cx == line_x + buffer is NOT right (it's in the zone)."""
        assert get_side(cx=310, line_x=300, buffer=10) is None

    def test_zero_buffer(self):
        assert get_side(cx=299, line_x=300, buffer=0) == "LEFT"
        assert get_side(cx=301, line_x=300, buffer=0) == "RIGHT"
        assert get_side(cx=300, line_x=300, buffer=0) is None


# ═══════════════════════════════════════════
# detect_crossing  (state-based)
# ═══════════════════════════════════════════

class TestDetectCrossing:
    """Test the state-based crossing-detection logic."""

    def _make_state(self):
        return {}, set(), set(), []

    def test_entry_left_to_right(self):
        """Person walks from left side to right side = ENTRY."""
        sides, entered, exited, log = self._make_state()
        # First sighting on left side
        detect_crossing(1, cx=280, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        assert sides[1] == "LEFT"
        assert len(log) == 0

        # Person walks into buffer zone — no change
        result = detect_crossing(1, cx=295, line_x=300, buffer=10,
                                 track_sides=sides, entered_ids=entered,
                                 exited_ids=exited, log_data=log)
        assert result is None
        assert sides[1] == "LEFT"  # still committed LEFT

        # Person emerges on right side — ENTRY!
        result = detect_crossing(1, cx=320, line_x=300, buffer=10,
                                 track_sides=sides, entered_ids=entered,
                                 exited_ids=exited, log_data=log)
        assert result == "ENTRY"
        assert 1 in entered
        assert len(log) == 1
        assert log[0][0] == 1        # track_id
        assert log[0][1] == "ENTRY"  # type

    def test_exit_right_to_left(self):
        """Person walks from right side to left side = EXIT."""
        sides, entered, exited, log = self._make_state()
        # First sighting on right
        detect_crossing(2, cx=320, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        # Into buffer zone
        detect_crossing(2, cx=305, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        # Emerge on left
        result = detect_crossing(2, cx=280, line_x=300, buffer=10,
                                 track_sides=sides, entered_ids=entered,
                                 exited_ids=exited, log_data=log)
        assert result == "EXIT"
        assert 2 in exited

    def test_slow_walk_across_many_frames(self):
        """Simulate a slow walk: 1-2 pixels per frame, crossing still detected."""
        sides, entered, exited, log = self._make_state()
        # Start on left, slowly walk right
        positions = list(range(270, 325, 2))  # 270, 272, 274, ... 324
        for cx in positions:
            detect_crossing(1, cx=cx, line_x=300, buffer=10,
                            track_sides=sides, entered_ids=entered,
                            exited_ids=exited, log_data=log)
        assert 1 in entered
        assert len(log) == 1
        assert log[0][0] == 1        # track_id
        assert log[0][1] == "ENTRY"  # type

    def test_no_crossing_stays_on_same_side(self):
        """Person stays on left side — no crossing."""
        sides, entered, exited, log = self._make_state()
        for cx in [250, 260, 270, 280, 285]:
            detect_crossing(3, cx=cx, line_x=300, buffer=10,
                            track_sides=sides, entered_ids=entered,
                            exited_ids=exited, log_data=log)
        assert len(log) == 0

    def test_no_crossing_stays_in_zone(self):
        """Person lingers in buffer zone — no crossing."""
        sides, entered, exited, log = self._make_state()
        # First sighting must be outside zone to establish side
        detect_crossing(4, cx=280, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        for cx in [292, 295, 300, 305, 298]:
            detect_crossing(4, cx=cx, line_x=300, buffer=10,
                            track_sides=sides, entered_ids=entered,
                            exited_ids=exited, log_data=log)
        # Side should still be LEFT since they never emerged on right
        assert sides[4] == "LEFT"
        assert len(log) == 0

    def test_duplicate_entry_ignored(self):
        """Same ID crossing a second time should NOT double-count."""
        sides, entered, exited, log = self._make_state()
        # First crossing
        detect_crossing(1, cx=280, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        detect_crossing(1, cx=320, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        assert len(log) == 1
        # Walk back and cross again
        detect_crossing(1, cx=280, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        detect_crossing(1, cx=320, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        # Still only 1 entry + 1 exit logged (not 2 entries)
        assert len(log) == 2  # 1 ENTRY + 1 EXIT

    def test_first_sighting_in_zone_no_crash(self):
        """If a person first appears inside the buffer zone, no side is set."""
        sides, entered, exited, log = self._make_state()
        result = detect_crossing(5, cx=300, line_x=300, buffer=10,
                                 track_sides=sides, entered_ids=entered,
                                 exited_ids=exited, log_data=log)
        assert result is None
        assert 5 not in sides

    def test_entry_and_exit_different_ids(self):
        """Two different people, one entering and one exiting."""
        sides, entered, exited, log = self._make_state()
        # Person 10: left → right
        detect_crossing(10, cx=280, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        detect_crossing(10, cx=320, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        # Person 20: right → left
        detect_crossing(20, cx=320, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        detect_crossing(20, cx=280, line_x=300, buffer=10,
                        track_sides=sides, entered_ids=entered,
                        exited_ids=exited, log_data=log)
        assert len(log) == 2
        assert log[0][0] == 10       # track_id for entry
        assert log[0][1] == "ENTRY"
        assert log[1][0] == 20       # track_id for exit
        assert log[1][1] == "EXIT"


# ═══════════════════════════════════════════
# prune_stale_tracks
# ═══════════════════════════════════════════

class TestPruneStaleTracks:
    """Test that stale (no longer visible) tracks are cleaned up."""

    def test_removes_stale(self):
        history = {1: "LEFT", 2: "RIGHT", 3: "LEFT"}
        prune_stale_tracks(history, active_ids={1, 3})
        assert 2 not in history
        assert len(history) == 2

    def test_keeps_all_active(self):
        history = {1: "LEFT", 2: "RIGHT"}
        prune_stale_tracks(history, active_ids={1, 2})
        assert len(history) == 2

    def test_empty_active_clears_all(self):
        history = {1: "LEFT", 2: "RIGHT"}
        prune_stale_tracks(history, active_ids=set())
        assert len(history) == 0

    def test_empty_history_is_noop(self):
        history = {}
        prune_stale_tracks(history, active_ids={1, 2, 3})
        assert len(history) == 0


# ═══════════════════════════════════════════
# handle_key
# ═══════════════════════════════════════════

class TestHandleKey:
    """Test keyboard input handling."""

    def test_esc_exits(self):
        should_exit, _, _, _ = handle_key(
            key=27, line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert should_exit is True

    @patch("counter.cv2")
    def test_f_toggles_fullscreen(self, mock_cv2):
        _, _, fs, _ = handle_key(
            key=ord('f'), line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert fs is True

    @patch("counter.cv2")
    def test_F_toggles_fullscreen(self, mock_cv2):
        _, _, fs, _ = handle_key(
            key=ord('F'), line_x=300, line_speed=10,
            frame_width=640, fullscreen=True, window_name="test",
        )
        assert fs is False

    def test_left_arrow_moves_line(self):
        _, lx, _, _ = handle_key(
            key=2424832, line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert lx == 290

    def test_right_arrow_moves_line(self):
        _, lx, _, _ = handle_key(
            key=2555904, line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert lx == 310

    def test_a_key_moves_line_left(self):
        _, lx, _, _ = handle_key(
            key=ord('a'), line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert lx == 290

    def test_d_key_moves_line_right(self):
        _, lx, _, _ = handle_key(
            key=ord('d'), line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert lx == 310

    def test_line_clamped_at_zero(self):
        _, lx, _, _ = handle_key(
            key=2424832, line_x=5, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert lx == 0

    def test_line_clamped_at_max(self):
        _, lx, _, _ = handle_key(
            key=2555904, line_x=635, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert lx == 639

    def test_unknown_key_no_change(self):
        should_exit, lx, fs, reset = handle_key(
            key=ord('z'), line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert should_exit is False
        assert lx == 300
        assert fs is False
        assert reset is False

    def test_r_key_triggers_reset(self):
        should_exit, _, _, reset = handle_key(
            key=ord('r'), line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert should_exit is False
        assert reset is True

    def test_R_key_triggers_reset(self):
        should_exit, _, _, reset = handle_key(
            key=ord('R'), line_x=300, line_speed=10,
            frame_width=640, fullscreen=False, window_name="test",
        )
        assert should_exit is False
        assert reset is True


# ═══════════════════════════════════════════
# generate_reports
# ═══════════════════════════════════════════

class TestGenerateReports:
    """Test CSV + Excel report generation."""

    def test_creates_files_with_data(self, tmp_path):
        csv_path = str(tmp_path / "log.csv")
        xlsx_path = str(tmp_path / "report.xlsx")
        xml_path = str(tmp_path / "log.xml")
        log_data = [
            [1, "ENTRY", "2026-04-24 12:00:00"],
            [2, "EXIT", "2026-04-24 12:01:00"],
            [3, "ENTRY", "2026-04-24 12:02:00"],
        ]
        generate_reports(log_data, csv_path=csv_path, excel_path=xlsx_path,
                         xml_path=xml_path)

        assert os.path.exists(csv_path)
        assert os.path.exists(xlsx_path)
        assert os.path.exists(xml_path)

        import pandas as pd
        df = pd.read_csv(csv_path)
        assert len(df) == 3
        assert list(df.columns) == ["TrackID", "Type", "Time"]

    def test_creates_files_empty_data(self, tmp_path):
        """Should not crash when there are zero crossings."""
        csv_path = str(tmp_path / "log.csv")
        xlsx_path = str(tmp_path / "report.xlsx")
        xml_path = str(tmp_path / "log.xml")
        generate_reports([], csv_path=csv_path, excel_path=xlsx_path,
                         xml_path=xml_path)

        assert os.path.exists(csv_path)
        assert os.path.exists(xlsx_path)
        assert os.path.exists(xml_path)

    def test_excel_summary_values(self, tmp_path):
        """Verify the summary cells in the Excel file."""
        xlsx_path = str(tmp_path / "report.xlsx")
        csv_path = str(tmp_path / "log.csv")
        xml_path = str(tmp_path / "log.xml")
        log_data = [
            [1, "ENTRY", "2026-04-24 12:00:00"],
            [2, "ENTRY", "2026-04-24 12:01:00"],
            [3, "EXIT", "2026-04-24 12:02:00"],
        ]
        generate_reports(log_data, csv_path=csv_path, excel_path=xlsx_path,
                         xml_path=xml_path)

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws["F2"].value == 2   # 2 entries
        assert ws["F3"].value == 1   # 1 exit


# ═══════════════════════════════════════════
# init_capture
# ═══════════════════════════════════════════

class TestInitCapture:
    """Test video capture initialization."""

    @patch("counter.cv2.VideoCapture")
    def test_raises_on_failure(self, mock_cap_cls):
        mock_cap = MagicMock()
        mock_cap.isOpened.return_value = False
        mock_cap_cls.return_value = mock_cap

        with pytest.raises(RuntimeError, match="Cannot open video source"):
            init_capture(99)

    @patch("counter.cv2.VideoCapture")
    def test_returns_capture_on_success(self, mock_cap_cls):
        mock_cap = MagicMock()
        mock_cap.isOpened.return_value = True
        mock_cap_cls.return_value = mock_cap

        result = init_capture(0)
        assert result is mock_cap


# ═══════════════════════════════════════════
# generate_xml
# ═══════════════════════════════════════════

class TestGenerateXml:
    """Test XML report generation."""

    def test_creates_xml_with_data(self, tmp_path):
        xml_path = str(tmp_path / "log.xml")
        log_data = [
            [1, "ENTRY", "2026-04-24 12:00:00"],
            [2, "EXIT", "2026-04-24 12:01:00"],
        ]
        generate_xml(log_data, xml_path=xml_path)
        assert os.path.exists(xml_path)

        tree = ET.parse(xml_path)
        root = tree.getroot()
        assert root.tag == "PeopleLog"

        summary = root.find("Summary")
        assert summary.find("TotalEntries").text == "1"
        assert summary.find("TotalExits").text == "1"

        events = root.find("Events")
        event_list = events.findall("Event")
        assert len(event_list) == 2
        assert event_list[0].find("TrackID").text == "1"
        assert event_list[0].find("Type").text == "ENTRY"
        assert event_list[1].find("TrackID").text == "2"
        assert event_list[1].find("Type").text == "EXIT"

    def test_creates_xml_empty_data(self, tmp_path):
        xml_path = str(tmp_path / "log.xml")
        generate_xml([], xml_path=xml_path)
        assert os.path.exists(xml_path)

        tree = ET.parse(xml_path)
        root = tree.getroot()
        assert root.find("Summary").find("TotalEntries").text == "0"
        assert root.find("Summary").find("TotalExits").text == "0"
        assert len(root.find("Events").findall("Event")) == 0


# ═══════════════════════════════════════════
# reset_counter
# ═══════════════════════════════════════════

class TestResetCounter:
    """Test the reset functionality."""

    def test_reset_clears_all_state(self, tmp_path, monkeypatch):
        """After reset, all tracking state should be empty."""
        # Point reports to tmp so they don't overwrite real files
        monkeypatch.setattr("counter.CSV_PATH", str(tmp_path / "log.csv"))
        monkeypatch.setattr("counter.EXCEL_PATH", str(tmp_path / "report.xlsx"))
        monkeypatch.setattr("counter.XML_PATH", str(tmp_path / "log.xml"))

        track_sides = {1: "LEFT", 2: "RIGHT"}
        entered_ids = {1}
        exited_ids = {2}
        log_data = [[1, "ENTRY", "2026-04-24 12:00:00"],
                     [2, "EXIT", "2026-04-24 12:01:00"]]

        reset_counter(track_sides, entered_ids, exited_ids, log_data)

        assert len(track_sides) == 0
        assert len(entered_ids) == 0
        assert len(exited_ids) == 0
        assert len(log_data) == 0

    def test_reset_saves_reports(self, tmp_path, monkeypatch):
        """Reset should generate CSV/Excel/XML before clearing."""
        csv_path = str(tmp_path / "log.csv")
        xlsx_path = str(tmp_path / "report.xlsx")
        xml_path = str(tmp_path / "log.xml")
        monkeypatch.setattr("counter.CSV_PATH", csv_path)
        monkeypatch.setattr("counter.EXCEL_PATH", xlsx_path)
        monkeypatch.setattr("counter.XML_PATH", xml_path)

        log_data = [[1, "ENTRY", "2026-04-24 12:00:00"]]
        reset_counter({}, {1}, set(), log_data)

        assert os.path.exists(csv_path)
        assert os.path.exists(xlsx_path)
        assert os.path.exists(xml_path)

    def test_reset_with_empty_data_no_crash(self, tmp_path, monkeypatch):
        """Reset with no data should not crash or create files."""
        csv_path = str(tmp_path / "log.csv")
        monkeypatch.setattr("counter.CSV_PATH", csv_path)

        reset_counter({}, set(), set(), [])
        # No files created when there's no data
        assert not os.path.exists(csv_path)
