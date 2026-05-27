"""
People Counter — Real-time entry/exit counter using YOLOv8 + OpenCV.

Uses a movable vertical line to detect people crossing left→right (ENTRY)
or right→left (EXIT). Generates CSV, XML, and Excel reports on reset and exit.
"""

from ultralytics import YOLO
import cv2
import numpy as np
import pandas as pd
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
# -------------------------
# SETTINGS
# -------------------------
MODEL_PATH = "yolov8n.pt"
VIDEO_SOURCE = 0        # webcam (or path to video file)
LINE_X_INIT = 300       # initial vertical line position
BUFFER = 10             # crossing hysteresis to avoid false triggers
LINE_SPEED = 10         # arrow key movement speed (pixels)
CSV_PATH = "people_log.csv"
EXCEL_PATH = "people_report.xlsx"
XML_PATH = "people_log.xml"
# -------------------------
# UI COLORS (BGR)
# -------------------------
COLOR_ENTRY      = (0, 220, 100)     # green
COLOR_EXIT       = (80, 100, 255)    # red-ish
COLOR_LINE       = (255, 180, 0)     # cyan-blue line
COLOR_LINE_ZONE  = (255, 180, 0, 40) # translucent zone
COLOR_BBOX       = (0, 255, 200)     # yellow-green box
COLOR_CENTER     = (0, 140, 255)     # orange center dot
COLOR_ID_TEXT    = (255, 255, 255)   # white ID label
COLOR_HUD_BG     = (30, 30, 30)      # dark HUD background
COLOR_FPS        = (0, 230, 255)     # amber FPS text
COLOR_HINT       = (180, 180, 180)   # grey hint text

def init_capture(source):
    """Open a video capture and verify it works.
    Args:
        source: Camera index (int) or video file path (str).
    Returns:
        cv2.VideoCapture object.
    Raises:
        RuntimeError: If the video source cannot be opened.
    """
    cap = cv2.VideoCapture(source)
    if not cap.isOpened():
        raise RuntimeError(
            f"Cannot open video source: {source}. "
            "Check that your camera is connected or the file path is correct."
        )
    return cap

def get_side(cx, line_x, buffer):
    """Determine which side of the counting line a point is on.
    Args:
        cx:     X-coordinate of the point.
        line_x: X-coordinate of the counting line.
        buffer: Dead-zone width on each side of the line.

    Returns:
        "LEFT", "RIGHT", or None if inside the buffer zone.
    """
    if cx < line_x - buffer:
        return "LEFT"
    elif cx > line_x + buffer:
        return "RIGHT"
    return None  # inside dead zone — side unchanged


def detect_crossing(track_id, cx, line_x, buffer,
                    track_sides, entered_ids, exited_ids, log_data):
    """State-based crossing detection.

    Tracks which "committed side" (LEFT or RIGHT) each person is on.
    The side only updates when the person is clearly past the buffer
    zone, so slow walking across the line is detected correctly.

    A crossing is registered when the committed side transitions:
      - LEFT → RIGHT  =  ENTRY
      - RIGHT → LEFT  =  EXIT

    Args:
        track_id:    Integer ID assigned by the tracker.
        cx:          Current center-x of the bounding box.
        line_x:      X-coordinate of the counting line.
        buffer:      Hysteresis buffer (pixels) on each side of the line.
        track_sides: Dict mapping track_id → last committed side ("LEFT"/"RIGHT").
        entered_ids: Set of IDs that have already been counted as entries.
        exited_ids:  Set of IDs that have already been counted as exits.
        log_data:    List to which [direction, timestamp] rows are appended.

    Returns:
        str or None — "ENTRY", "EXIT", or None if no crossing detected.
    """
    current_side = get_side(cx, line_x, buffer)

    # Inside the dead zone — don't change committed side
    if current_side is None:
        return None

    prev_side = track_sides.get(track_id)

    # First sighting: record side, no crossing yet
    if prev_side is None:
        track_sides[track_id] = current_side
        return None

    # Side hasn't changed
    if current_side == prev_side:
        return None

    # Side changed — crossing detected!
    track_sides[track_id] = current_side

    # LEFT → RIGHT = ENTRY
    if prev_side == "LEFT" and current_side == "RIGHT":
        if track_id not in entered_ids:
            entered_ids.add(track_id)
            log_data.append([track_id, "ENTRY", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            return "ENTRY"

    # RIGHT → LEFT = EXIT
    elif prev_side == "RIGHT" and current_side == "LEFT":
        if track_id not in exited_ids:
            exited_ids.add(track_id)
            log_data.append([track_id, "EXIT", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            return "EXIT"

    return None


def prune_stale_tracks(track_history, active_ids):
    """Remove track IDs that are no longer detected.

    Prevents the track_history dict from growing unbounded when objects
    leave the scene and their IDs are recycled.

    Args:
        track_history: Dict mapping track_id → last known data.
        active_ids:    Set/list of currently visible track IDs.
    """
    stale = [tid for tid in track_history if tid not in active_ids]
    for tid in stale:
        del track_history[tid]


def draw_overlay(frame, alpha=0.55):
    """Create a semi-transparent overlay copy for blending HUD elements.

    Args:
        frame: BGR image (numpy array).
        alpha: Opacity for the overlay blend (0=invisible, 1=opaque).

    Returns:
        (overlay, alpha) — overlay is a copy of frame to draw on.
    """
    return frame.copy(), alpha


def draw_hud_panel(frame, x, y, w, h, color=COLOR_HUD_BG, alpha=0.6):
    """Draw a rounded semi-transparent rectangle as a HUD background.

    Args:
        frame: BGR image to draw on (modified in-place).
        x, y:  Top-left corner.
        w, h:  Width and height.
        color: BGR fill color.
        alpha: Transparency (0=transparent, 1=opaque).
    """
    overlay = frame.copy()
    # Draw filled rounded rectangle (approximated with rectangle + circles)
    radius = min(12, h // 2, w // 2)
    # Main body
    cv2.rectangle(overlay, (x + radius, y), (x + w - radius, y + h), color, -1)
    cv2.rectangle(overlay, (x, y + radius), (x + w, y + h - radius), color, -1)
    # Four corners
    cv2.circle(overlay, (x + radius, y + radius), radius, color, -1)
    cv2.circle(overlay, (x + w - radius, y + radius), radius, color, -1)
    cv2.circle(overlay, (x + radius, y + h - radius), radius, color, -1)
    cv2.circle(overlay, (x + w - radius, y + h - radius), radius, color, -1)
    cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)


def draw_counting_line(frame, line_x, buffer, h):
    """Draw the counting line with a translucent buffer zone.

    Args:
        frame:  BGR image to draw on (modified in-place).
        line_x: X-coordinate of the center line.
        buffer: Width of the hysteresis zone on each side.
        h:      Frame height.
    """
    # Buffer zone (translucent strip)
    zone_overlay = frame.copy()
    left = max(0, line_x - buffer)
    right = min(frame.shape[1], line_x + buffer)
    cv2.rectangle(zone_overlay, (left, 0), (right, h), COLOR_LINE, -1)
    cv2.addWeighted(zone_overlay, 0.12, frame, 0.88, 0, frame)

    # Main line — solid with glow effect
    cv2.line(frame, (line_x, 0), (line_x, h), COLOR_LINE, 3)
    # Thin bright core
    cv2.line(frame, (line_x, 0), (line_x, h), (255, 255, 255), 1)

    # Arrow indicators at top/bottom of line
    arrow_sz = 8
    # Top: entry arrow (→)
    cv2.arrowedLine(frame, (line_x - 20, 25), (line_x + 20, 25),
                    COLOR_ENTRY, 2, tipLength=0.4)
    cv2.putText(frame, "IN", (line_x + 24, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.4, COLOR_ENTRY, 1)
    # Bottom: exit arrow (←)
    cv2.arrowedLine(frame, (line_x + 20, h - 45), (line_x - 20, h - 45),
                    COLOR_EXIT, 2, tipLength=0.4)
    cv2.putText(frame, "OUT", (line_x - 50, h - 40),
                cv2.FONT_HERSHEY_SIMPLEX, 0.4, COLOR_EXIT, 1)


def draw_bbox(frame, x1, y1, x2, y2, track_id, cx, cy, line_x, buffer):
    """Draw a styled bounding box with ID label and center dot.

    Color changes based on which side of the line the person is on.

    Args:
        frame:    BGR image to draw on.
        x1..y2:   Bounding box coordinates.
        track_id: Tracker ID for labeling.
        cx, cy:   Center point of the box.
        line_x:   Counting line X position.
        buffer:   Buffer zone width.
    """
    # Color based on position relative to line
    if cx < line_x - buffer:
        box_color = COLOR_EXIT       # left side = potential exit direction
    elif cx > line_x + buffer:
        box_color = COLOR_ENTRY      # right side = already entered
    else:
        box_color = COLOR_LINE       # in the zone

    # Stylish corner-only bounding box
    corner_len = max(15, min((x2 - x1), (y2 - y1)) // 4)
    thickness = 2

    # Top-left
    cv2.line(frame, (x1, y1), (x1 + corner_len, y1), box_color, thickness)
    cv2.line(frame, (x1, y1), (x1, y1 + corner_len), box_color, thickness)
    # Top-right
    cv2.line(frame, (x2, y1), (x2 - corner_len, y1), box_color, thickness)
    cv2.line(frame, (x2, y1), (x2, y1 + corner_len), box_color, thickness)
    # Bottom-left
    cv2.line(frame, (x1, y2), (x1 + corner_len, y2), box_color, thickness)
    cv2.line(frame, (x1, y2), (x1, y2 - corner_len), box_color, thickness)
    # Bottom-right
    cv2.line(frame, (x2, y2), (x2 - corner_len, y2), box_color, thickness)
    cv2.line(frame, (x2, y2), (x2, y2 - corner_len), box_color, thickness)

    # Center dot with ring
    cv2.circle(frame, (cx, cy), 6, box_color, 1)
    cv2.circle(frame, (cx, cy), 2, COLOR_CENTER, -1)

    # ID label with background
    label = f"ID {track_id}"
    (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.45, 1)
    label_x, label_y = x1, y1 - 8
    cv2.rectangle(frame, (label_x - 1, label_y - th - 4),
                  (label_x + tw + 4, label_y + 4), box_color, -1)
    cv2.putText(frame, label, (label_x + 2, label_y),
                cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 0), 1)


def process_frame(frame, model, line_x, buffer,
                  track_sides, entered_ids, exited_ids, log_data,
                  fps=0.0):
    """Run detection + tracking on a single frame and annotate it.

    Args:
        frame:        BGR image (numpy array).
        model:        YOLO model instance.
        line_x:       X-coordinate of the counting line.
        buffer:       Crossing hysteresis buffer.
        track_sides:  Dict mapping track_id → committed side ("LEFT"/"RIGHT").
        entered_ids:  Set of IDs counted as entered.
        exited_ids:   Set of IDs counted as exited.
        log_data:     List collecting [direction, timestamp] rows.
        fps:          Current frames-per-second to display.

    Returns:
        Annotated frame (numpy array).
    """
    results = model.track(frame, persist=True, classes=[0], verbose=False)

    h, w = frame.shape[:2]
    active_ids = set()
    person_count = 0

    if results and results[0].boxes.id is not None:
        boxes = results[0].boxes.xyxy
        ids = results[0].boxes.id
        person_count = len(ids)

        for box, tid in zip(boxes, ids):
            x1, y1, x2, y2 = map(int, box)
            track_id = int(tid)
            active_ids.add(track_id)

            cx = (x1 + x2) // 2
            cy = (y1 + y2) // 2

            # Draw styled bounding box
            draw_bbox(frame, x1, y1, x2, y2, track_id, cx, cy, line_x, buffer)

            # Detect crossing using state-based approach
            detect_crossing(track_id, cx, line_x, buffer,
                            track_sides, entered_ids, exited_ids, log_data)

    # Prune tracks that are no longer visible
    prune_stale_tracks(track_sides, active_ids)

    # Draw counting line with buffer zone
    draw_counting_line(frame, line_x, buffer, h)

    # ─── HUD: Top-left stats panel ───
    draw_hud_panel(frame, 10, 8, 200, 110)

    # FPS
    fps_color = (0, 255, 150) if fps >= 20 else COLOR_FPS if fps >= 10 else (0, 80, 255)
    cv2.putText(frame, f"FPS: {fps:.1f}", (22, 35),
                cv2.FONT_HERSHEY_SIMPLEX, 0.55, fps_color, 2)

    # Entry / Exit counts
    cv2.putText(frame, f"IN:  {len(entered_ids)}", (22, 65),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, COLOR_ENTRY, 2)
    cv2.putText(frame, f"OUT: {len(exited_ids)}", (22, 95),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, COLOR_EXIT, 2)

    # Person count badge (top-right)
    badge_text = f"Tracking: {person_count}"
    (btw, bth), _ = cv2.getTextSize(badge_text, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
    draw_hud_panel(frame, w - btw - 30, 8, btw + 20, bth + 16)
    cv2.putText(frame, badge_text, (w - btw - 20, 8 + bth + 6),
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

    # ─── HUD: Bottom controls hint ───
    hint = "LEFT/RIGHT: Move Line | F: Fullscreen | R: Reset | ESC: Exit"
    (htw, hth), _ = cv2.getTextSize(hint, cv2.FONT_HERSHEY_SIMPLEX, 0.4, 1)
    draw_hud_panel(frame, 10, h - hth - 22, htw + 16, hth + 14)
    cv2.putText(frame, hint, (18, h - 18),
                cv2.FONT_HERSHEY_SIMPLEX, 0.4, COLOR_HINT, 1)

    return frame


def handle_key(key, line_x, line_speed, frame_width, fullscreen, window_name):
    """Process a keyboard event and return updated state.

    Uses key codes from cv2.waitKeyEx() which returns full extended
    key codes on all platforms (including Windows arrow keys).

    Args:
        key:          Key code from cv2.waitKeyEx.
        line_x:       Current line X position.
        line_speed:   Pixels to move per key press.
        frame_width:  Width of the current frame.
        fullscreen:   Current fullscreen state (bool).
        window_name:  Name of the OpenCV window.

    Returns:
        Tuple (should_exit, line_x, fullscreen, should_reset).
    """
    should_exit = False
    should_reset = False

    if key == 27:  # ESC
        should_exit = True

    elif key in (ord('r'), ord('R')):
        should_reset = True

    elif key in (ord('f'), ord('F')):
        fullscreen = not fullscreen
        prop = cv2.WINDOW_FULLSCREEN if fullscreen else cv2.WINDOW_NORMAL
        cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, prop)

    # Arrow keys — waitKeyEx returns these on Windows
    elif key in (2424832, 0x250000, 81, 63234):  # LEFT arrow
        line_x -= line_speed
    elif key in (2555904, 0x270000, 83, 63235):  # RIGHT arrow
        line_x += line_speed

    # Also support A/D keys for line movement
    elif key in (ord('a'), ord('A')):
        line_x -= line_speed
    elif key in (ord('d'), ord('D')):
        line_x += line_speed

    # Clamp line to valid range
    line_x = max(0, min(line_x, frame_width - 1))

    return should_exit, line_x, fullscreen, should_reset


def generate_xml(log_data, xml_path=XML_PATH):
    """Save an XML log of all entry/exit events.

    Args:
        log_data: List of [track_id, direction, timestamp] rows.
        xml_path: Output path for the XML file.
    """
    root = ET.Element("PeopleLog")
    root.set("generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    entry_count = sum(1 for row in log_data if row[1] == "ENTRY")
    exit_count = sum(1 for row in log_data if row[1] == "EXIT")

    summary = ET.SubElement(root, "Summary")
    ET.SubElement(summary, "TotalEntries").text = str(entry_count)
    ET.SubElement(summary, "TotalExits").text = str(exit_count)

    events = ET.SubElement(root, "Events")
    for row in log_data:
        event = ET.SubElement(events, "Event")
        ET.SubElement(event, "TrackID").text = str(row[0])
        ET.SubElement(event, "Type").text = row[1]
        ET.SubElement(event, "Time").text = row[2]

    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")  # Pretty-print
    tree.write(xml_path, encoding="utf-8", xml_declaration=True)
    print(f"✅ XML report saved to {xml_path}")


def generate_reports(log_data, csv_path=CSV_PATH, excel_path=EXCEL_PATH,
                     xml_path=XML_PATH):
    """Save a CSV log, XML log, and an Excel report with a pie chart.

    Args:
        log_data:   List of [track_id, direction, timestamp] rows.
        csv_path:   Output path for the CSV file.
        excel_path: Output path for the Excel file.
        xml_path:   Output path for the XML file.
    """
    df = pd.DataFrame(log_data, columns=["TrackID", "Type", "Time"])
    df.to_csv(csv_path, index=False)

    entry_count = len(df[df["Type"] == "ENTRY"])
    exit_count = len(df[df["Type"] == "EXIT"])

    df.to_excel(excel_path, index=False)
    wb = load_workbook(excel_path)
    ws = wb.active

    # Write summary block
    ws["E1"] = "Summary"
    ws["E2"] = "Entry"
    ws["E3"] = "Exit"
    ws["F2"] = entry_count
    ws["F3"] = exit_count

    # Only add a pie chart if there is data to chart
    if entry_count + exit_count > 0:
        pie = PieChart()
        labels = Reference(ws, min_col=5, min_row=2, max_row=3)
        data = Reference(ws, min_col=6, min_row=2, max_row=3)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.title = "Entry vs Exit"
        ws.add_chart(pie, "H2")

    wb.save(excel_path)

    # Also generate XML report
    generate_xml(log_data, xml_path)
    print("✅ CSV + Excel + XML reports generated!")


def reset_counter(track_sides, entered_ids, exited_ids, log_data):
    """Save current data to CSV/XML/Excel then reset all counters.

    Args:
        track_sides: Dict mapping track_id → committed side.
        entered_ids: Set of IDs counted as entered.
        exited_ids:  Set of IDs counted as exited.
        log_data:    List of [track_id, direction, timestamp] rows.
    """
    if log_data:
        generate_reports(log_data, csv_path=CSV_PATH, excel_path=EXCEL_PATH,
                         xml_path=XML_PATH)
        print("🔄 Counter reset — reports saved.")
    else:
        print("🔄 Counter reset — no data to save.")

    track_sides.clear()
    entered_ids.clear()
    exited_ids.clear()
    log_data.clear()


def run_counter():
    """Main loop: capture frames, track people, count crossings."""
    model = YOLO(MODEL_PATH)
    cap = init_capture(VIDEO_SOURCE)

    track_sides = {}   # track_id → committed side ("LEFT" / "RIGHT")
    entered_ids = set()
    exited_ids = set()
    log_data = []

    window_name = "People Counter"
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    fullscreen = False
    line_x = LINE_X_INIT

    # FPS tracking
    prev_time = time.time()
    fps = 0.0

    try:
        while True:
            ret, frame = cap.read()
            if not ret:
                print("⚠️  No frame received — camera disconnected or video ended.")
                break

            # Calculate FPS
            curr_time = time.time()
            time_diff = curr_time - prev_time
            fps = 1.0 / time_diff if time_diff > 0 else 0.0
            prev_time = curr_time

            frame = process_frame(
                frame, model, line_x, BUFFER,
                track_sides, entered_ids, exited_ids, log_data,
                fps=fps
            )
            cv2.imshow(window_name, frame)

            # Use waitKeyEx for proper extended key codes (arrow keys on Windows)
            key = cv2.waitKeyEx(1)
            should_exit, line_x, fullscreen, should_reset = handle_key(
                key, line_x, LINE_SPEED, frame.shape[1],
                fullscreen, window_name
            )

            if should_reset:
                reset_counter(track_sides, entered_ids, exited_ids, log_data)

            if should_exit:
                break
    finally:
        cap.release()
        cv2.destroyAllWindows()

    # Generate reports after the window is closed (on exit)
    if log_data:
        generate_reports(log_data)


if __name__ == "__main__":
    run_counter()